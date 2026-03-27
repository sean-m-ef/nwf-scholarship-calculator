"""Excel output builder using openpyxl."""

from __future__ import annotations

import io

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from engine.postprocess import ProcessedResult

_YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
_RED_LIGHT = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
_HEADER_FONT = Font(bold=True)
_CURRENCY_FMT = '"$"#,##0.00'


def build_excel(pr: ProcessedResult) -> bytes:
    """Return an Excel workbook as bytes with three sheets."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    _write_allocations(wb, pr)
    _write_recipient_summary(wb, pr)
    _write_scholarship_summary(wb, pr)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_allocations(wb: openpyxl.Workbook, pr: ProcessedResult) -> None:
    ws = wb.create_sheet("Allocations")
    df = pr.allocations

    headers = ["Recipient ID", "Recipient Name", "Scholarship ID", "Scholarship Name", "Amount", "Small Split Warning"]
    col_keys = ["recipient_id", "recipient_name", "scholarship_id", "scholarship_name", "amount", "small_split"]

    _write_header(ws, headers)

    for _, row in df.iterrows():
        values = [row[k] for k in col_keys]
        values[-1] = "YES" if row["small_split"] else ""
        ws.append(values)
        r = ws.max_row
        # Currency format for amount column (col 5)
        ws.cell(r, 5).number_format = _CURRENCY_FMT
        if row["small_split"]:
            for c in range(1, len(headers) + 1):
                ws.cell(r, c).fill = _YELLOW

    _autosize(ws)


def _write_recipient_summary(wb: openpyxl.Workbook, pr: ProcessedResult) -> None:
    ws = wb.create_sheet("Recipient Summary")
    df = pr.recipient_summary

    headers = ["Recipient ID", "Name", "Award Amount", "Total Allocated", "Gap", "# Scholarships", "Small Split?", "Infeasible?"]
    col_keys = ["recipient_id", "name", "award_amount", "total_allocated", "gap", "num_scholarships", "has_small_split", "infeasible"]

    _write_header(ws, headers)

    for _, row in df.iterrows():
        values = [row[k] for k in col_keys]
        # Convert booleans to readable strings
        values[6] = "YES" if row["has_small_split"] else ""
        values[7] = "YES" if row["infeasible"] else ""
        ws.append(values)
        r = ws.max_row
        for c in (3, 4, 5):  # currency columns
            ws.cell(r, c).number_format = _CURRENCY_FMT
        if row["infeasible"]:
            for c in range(1, len(headers) + 1):
                ws.cell(r, c).fill = _RED_LIGHT
        elif row["has_small_split"]:
            for c in range(1, len(headers) + 1):
                ws.cell(r, c).fill = _YELLOW

    _autosize(ws)


def _write_scholarship_summary(wb: openpyxl.Workbook, pr: ProcessedResult) -> None:
    ws = wb.create_sheet("Scholarship Summary")
    df = pr.scholarship_summary

    headers = ["Scholarship ID", "Scholarship Name", "Total Available", "Total Disbursed", "Remaining", "# Recipients"]
    col_keys = ["scholarship_id", "scholarship_name", "total_available", "total_disbursed", "remaining", "num_recipients"]

    _write_header(ws, headers)

    for _, row in df.iterrows():
        values = [row[k] for k in col_keys]
        ws.append(values)
        r = ws.max_row
        for c in (3, 4, 5):  # currency columns
            ws.cell(r, c).number_format = _CURRENCY_FMT
        if row["remaining"] > 0.01:
            # Undisbursed scholarship — highlight for review
            for c in range(1, len(headers) + 1):
                ws.cell(r, c).fill = _YELLOW

    _autosize(ws)


def _write_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _autosize(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)
