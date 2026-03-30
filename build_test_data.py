"""Extract client Excel spreadsheet into engine-compatible CSVs for test run.

Reads /Users/sean/Downloads/NWF Scholarships to SM.xlsx and produces:
  test_data/recipients.csv      — funded recipients with attribute profile
  test_data/scholarships.csv    — scholarships with crit__ criteria columns
  test_data/manual_allocations.csv — client's manual allocation for comparison

Run: uv run python build_test_data.py
"""

from __future__ import annotations

import csv
import math
import os
from datetime import date
from pathlib import Path

import openpyxl

EXCEL_PATH = Path("/Users/sean/Downloads/NWF Scholarships to SM.xlsx")
OUT_DIR = Path("test_data")

# Cities considered urban/non-rural in Northern Nevada
URBAN_CITIES = {
    "reno", "sparks", "sparks ", "carson city", "sun valley",
    "incline village", "verdi",
}

# Douglas County cities
DOUGLAS_COUNTY_CITIES = {"minden", "gardnerville", "genoa", "stateline"}

# Cities qualifying for 180 Gold (Winnemucca / Battle Mountain / Eureka)
NORTHERN_RURAL_CITIES = {"winnemucca", "battle mountain", "eureka"}

# Washoe County cities (for Redfield scholarship)
WASHOE_COUNTY_CITIES = {"reno", "sparks", "sparks ", "sun valley", "verdi"}

# STEM major keywords (used to derive stem column since spreadsheet field is empty)
STEM_KEYWORDS = [
    "engineering", "computer science", "mathematics", "biochemistry",
    "chemistry", "physics", "biology", "data science", "metallurgical",
    "radiologic", "sonography", "technology", "ecology", "science",
    "medicine", "veterinary", "pre-nursing", "nursing",
]

# School name normalization: substring → canonical key
SCHOOL_MAP = [
    ("university of nevada, reno", "unr"),
    ("truckee meadows", "tmcc"),
    ("western nevada college", "wnc"),
    ("great basin college", "gbc"),
    ("western governor", "wgu"),
    ("unlv", "unlv"),
    ("uunlv", "unlv"),
    ("boyd", "unlv"),   # Boyd = UNLV law
]

REFERENCE_DATE = date(2025, 9, 1)  # Fall 2025 semester start


def normalize_school(raw: str | None) -> str:
    """Scan raw school string for known patterns; multiple matches become comma-joined keys.

    Important: do NOT split on comma before matching — "University of Nevada, Reno"
    contains a comma and would be destroyed by a naïve split.
    """
    if not raw:
        return "other"
    raw_lower = raw.lower()
    keys = []
    for substr, key in SCHOOL_MAP:
        if substr in raw_lower and key not in keys:
            keys.append(key)
    return ",".join(keys) if keys else "other"


def normalize_grade(raw: str | None) -> str:
    if not raw:
        return "other"
    r = raw.strip().lower()
    if r in ("freshman", "freshman "):
        return "freshman"
    if r in ("sophomore", ):
        return "sophomore"
    if r in ("junior", ):
        return "junior"
    if r in ("senior", ):
        return "senior"
    if r in ("graduate", ):
        return "graduate"
    if r in ("doctorate", ):
        return "doctorate"
    if r in ("wr", ):
        return "wr"
    return "other"


def normalize_degree(raw: str | None) -> str:
    if not raw:
        return "other"
    r = raw.strip().lower()
    if "bachelor" in r:
        return "bachelor"
    if "master" in r:
        return "masters"
    if "doctor of medicine" in r or "md" in r:
        return "md"
    if "doctor of veterinary" in r or "dvm" in r:
        return "dvm"
    if "juris doctor" in r or "jd" in r:
        return "jd"
    if "doctorate" in r:
        return "doctorate"
    if "associate" in r:
        return "associate"
    if "certification" in r or "professional" in r:
        return "professional"
    if "trade" in r or "vocational" in r:
        return "trade"
    if "community college" in r or "continuing education" in r:
        return "associate"
    return "other"


def infer_stem(major: str | None) -> str:
    if not major:
        return "no"
    m = major.lower()
    for kw in STEM_KEYWORDS:
        if kw in m:
            return "yes"
    return "no"


def yn(val) -> str:
    """Normalize Yes/No/None to yes/no."""
    if val is None:
        return "no"
    if str(val).strip().lower() in ("yes", "true", "1"):
        return "yes"
    return "no"


def compute_age(birth) -> int | None:
    if birth is None:
        return None
    if hasattr(birth, "date"):
        birth = birth.date()
    delta = REFERENCE_DATE - birth
    return delta.days // 365


def build_recipients(ws) -> tuple[list[dict], dict[int, dict]]:
    """Return (recipients_rows, col_to_scholarship) where col_to_scholarship
    maps Excel column index → scholarship dict for building manual allocations."""

    # Read scholarship metadata from header rows
    scholarships_by_col: dict[int, dict] = {}
    for col in range(42, 218):
        name = ws.cell(2, col).value
        if name and str(name).strip():
            scholarships_by_col[col] = {"name": str(name).strip()}

    recipients = []
    # Collect all data rows first to assign clean IDs
    raw_rows = []
    for row in ws.iter_rows(min_row=7, max_row=242, values_only=True):
        num = row[1]
        if num is None or str(num).strip() in ("", " "):
            continue
        actual_amount = row[39]
        if actual_amount is None or not isinstance(actual_amount, (int, float)):
            continue
        if actual_amount <= 0:
            continue
        raw_rows.append(row)

    for idx, row in enumerate(raw_rows):
        num = row[1]
        rid = f"R{str(num).strip().zfill(3)}"

        ethnicity = str(row[2]).strip().lower() if row[2] else ""
        marital_status = str(row[3]).strip().lower() if row[3] else ""
        age = compute_age(row[4])
        degree = normalize_degree(row[5])
        city_raw = str(row[6]).strip() if row[6] else ""
        city = city_raw.lower()
        major_raw = str(row[7]).strip() if row[7] else ""
        major = major_raw.lower()
        grade = normalize_grade(row[11])
        school = normalize_school(row[12])
        gap_in_studies = yn(row[15])
        has_minor_dependents = yn(row[22])
        # Fall back to number_of_children if dependents question unanswered
        if has_minor_dependents == "no" and row[22] is None:
            n_children = row[18]
            if n_children and str(n_children).strip() not in ("", " ", "None"):
                try:
                    if int(str(n_children).strip()) > 0:
                        has_minor_dependents = "yes"
                except (ValueError, TypeError):
                    pass
        parents_graduated = yn(row[24])
        construction_field = yn(row[25])
        silversummit = yn(row[32])
        deans_future_scholars = yn(row[33])
        tutoring = yn(row[34])
        award_amount = float(row[39])

        # Derived fields
        first_gen = "yes" if parents_graduated == "no" else "no"
        rural = "no" if city in URBAN_CITIES else "yes"
        douglas_county = "yes" if city in DOUGLAS_COUNTY_CITIES else "no"
        washoe_county = "yes" if city in WASHOE_COUNTY_CITIES else "no"
        northern_rural = "yes" if city in NORTHERN_RURAL_CITIES else "no"
        single_mother = "yes" if (
            marital_status in ("single", "divorced") and has_minor_dependents == "yes"
        ) else "no"
        stem = infer_stem(major)

        recipients.append({
            "recipient_id": rid,
            "full_name": f"Recipient {str(num).strip()}",
            "award_amount": round(award_amount, 2),
            "ethnicity": ethnicity,
            "marital_status": marital_status,
            "age": age if age is not None else "",
            "city": city,
            "school": school,
            "major": major,
            "degree": degree,
            "grade": grade,
            "gap_in_studies": gap_in_studies,
            "silversummit": silversummit,
            "construction_field": construction_field,
            "tutoring": tutoring,
            "deans_future_scholars": deans_future_scholars,
            "has_children": has_minor_dependents,
            "parents_graduated": parents_graduated,
            "first_gen": first_gen,
            "rural": rural,
            "douglas_county": douglas_county,
            "washoe_county": washoe_county,
            "northern_rural": northern_rural,
            "single_mother": single_mother,
            "stem": stem,
            # Store row for allocation extraction
            "_row": row,
        })

    return recipients, scholarships_by_col


def build_scholarships() -> list[dict]:
    """Return scholarship rows with crit__ criteria columns.

    Criteria notes:
    - OR logic is not supported; scholarships with OR criteria use the
      most restrictive single criterion or are left general.
    - Data gaps (AHEC membership, domestic violence, incarceration,
      Soroptimist membership) are treated as general (no criteria).
    - 'stem' is derived from major keywords, not from spreadsheet field
      (which was empty for all recipients).
    """

    def row(**kwargs) -> dict:
        return kwargs

    # Unique criteria column names across all scholarships
    # Format: crit__{attribute}__{operator}__{value}
    CRIT_COLS = [
        "crit__school__contains__unr",
        "crit__school__contains__tmcc",
        "crit__school__contains__wnc",
        "crit__school__contains__gbc",
        "crit__major__contains__nursing",
        "crit__major__contains__education",
        "crit__major__contains__law",
        "crit__major__contains__engineering",
        "crit__major__contains__medicine",
        "crit__major__contains__business",
        "crit__major__contains__accounting",
        "crit__major__contains__construction",
        "crit__major__contains__veterinary",
        "crit__major__contains__music",
        "crit__major__contains__black",
        "crit__ethnicity__contains__black",
        "crit__gap_in_studies__eq__yes",
        "crit__silversummit__eq__yes",
        "crit__construction_field__eq__yes",
        "crit__tutoring__eq__yes",
        "crit__first_gen__eq__yes",
        "crit__rural__eq__yes",
        "crit__douglas_county__eq__yes",
        "crit__washoe_county__eq__yes",
        "crit__northern_rural__eq__yes",
        "crit__single_mother__eq__yes",
        "crit__stem__eq__yes",
        "crit__age__gte__40",
        "crit__age__gte__25",
        "crit__age__lte__55",
        "crit__grade__eq__freshman",
        "crit__city__eq__yerington",
    ]

    def s(scholarship_id, name, amount, **active_crits) -> dict:
        """Build a scholarship row with all crit columns defaulting to false."""
        d = {
            "scholarship_id": scholarship_id,
            "name": name,
            "amount": amount,
        }
        for col in CRIT_COLS:
            key = col  # column name is the crit__ key
            d[col] = "true" if active_crits.get(key) else "false"
        return d

    T = True  # shorthand

    return [
        s("S001", "AAUW Capital", 6000,
          **{"crit__school__contains__wnc": T, "crit__gap_in_studies__eq__yes": T}),
        s("S002", "AAUW Reno", 5000),   # general (re-entry + career change, no clean map)
        s("S003", "Access To Healthcare Degree", 2500),   # AHEC — no data
        s("S004", "Access to Healthcare Non-Degree", 2400),   # AHEC — no data
        s("S005", "AHEC Degree", 12500),   # AHEC — no data
        s("S006", "AHEC Non-degree", 2400),   # AHEC — no data
        s("S007", "Amanda Fisher Endowed", 482.67),   # incarceration — no data
        s("S008", "Amber Planeta", 2500),
        s("S009", "Amy Biehl", 2000,
          **{"crit__ethnicity__contains__black": T}),
        s("S010", "ATT Endowed", 492.67),
        s("S011", "Barbara Nielsen", 464.69),   # art or music — OR logic, general
        s("S012", "Beverly Jean Acton Endowed", 773.28),
        s("S013", "Bill&Dottie Raggio Endowed", 3779.44),   # public service — general
        s("S014", "Bill&Moya Lear Endowed", 1145.56),
        s("S015", "Brookfield School Annual", 3000,
          **{"crit__major__contains__education": T}),
        s("S016", "Capital Insurance Annual", 3000),   # tbd
        s("S017", "Charles H. Stout Endowed", 2702.91,
          **{"crit__gap_in_studies__eq__yes": T}),
        s("S018", "Charles River", 2500),   # sci/vet/lab — OR logic, general
        s("S019", "Charlotte L. MacKenzie Endowed", 6508.04),   # full-time = general
        s("S020", "Courage Sans Peur Annual", 10000),   # Northern NV = general
        s("S021", "Davis Family Endowed", 3576.86,
          **{"crit__major__contains__nursing": T}),
        s("S022", "Dena L. Kruse Annual", 5000),   # health sciences — OR w/ nursing, general
        s("S023", "Derrivan/Rinaldi Endowed", 647.2,
          **{"crit__major__contains__music": T}),
        s("S024", "Dorothy Blakey Endowed", 6955.2),
        s("S025", "Dr. Kristen McNeil Endowed", 576.38,
          **{"crit__major__contains__education": T}),
        s("S026", "Dr. Rita Huneycutt Endowed", 2573.64,
          **{"crit__school__contains__tmcc": T}),
        s("S027", "Dr. Sandra A. Daugherty Endowed", 1243.39,
          **{"crit__school__contains__unr": T, "crit__major__contains__medicine": T}),
        s("S028", "E.L. Cord", 20000),
        s("S029", "Empowering Women of Douglas County", 3930,
          **{"crit__douglas_county__eq__yes": T}),
        s("S030", "Empowering Women of Yerington", 3075,
          **{"crit__city__eq__yerington": T}),
        s("S031", "Empowerment Ventures - Hindi", 2500),   # entrepreneurship — general
        s("S032", "EWB Math and Science Endowed", 15000,
          **{"crit__stem__eq__yes": T}),
        s("S033", "Ezell", 2500,
          **{"crit__age__gte__40": T}),
        s("S034", "Faralie S. Spell Memorial Endowed", 1566.98),   # education OR Native American — general
        s("S035", "Frankie Sue Del Papa Endowed", 2300.88),   # poli sci/student gov — no data, general
        s("S036", "Fritsi H. Ericson Endowed", 5000),
        s("S037", "Georgia Loy", 2500),
        s("S038", "Gonsalez/Morow Group", 2500),
        s("S039", "Guild Family Law Endowed", 1241.83,
          **{"crit__major__contains__law": T}),
        s("S040", "180 Gold", 5000,
          **{"crit__northern_rural__eq__yes": T}),
        s("S041", "IGT", 7500,
          **{"crit__stem__eq__yes": T}),
        s("S042", "Jan Evans Endowed", 3623.61,
          **{"crit__single_mother__eq__yes": T}),   # partial: also requires education/social work
        s("S043", "Jay Family Scholarship", 2500),
        s("S044", "Jennifer Satre Endowed", 2500),
        s("S045", "Joan & Robert Dees Endowed", 1000,
          **{"crit__major__contains__nursing": T}),
        s("S046", "Joan Dees Endowed", 2000,
          **{"crit__major__contains__nursing": T}),
        s("S047", "June Wisham Memorial Endowed", 28040.62),
        s("S048", "Kate Hanlon", 14000),
        s("S049", "Laura Dianda Endowed", 1063.55,
          **{"crit__construction_field__eq__yes": T}),
        s("S050", "Louise Jacobsen", 2000),   # nursing/science/math/arts — OR, general
        s("S051", "Maggie McGrew", 2000),   # tbd
        s("S052", "Marco/Schaffner Endowed", 703.09,
          **{"crit__school__contains__unr": T, "crit__major__contains__business": T}),
        s("S053", "Margaret Eddelman O'Donnell Endowed", 1368.83,
          **{"crit__school__contains__unr": T}),   # UNR Junior+ — grade OR not expressible
        s("S054", "Margaret Hart Conaboy Endowed", 2500,
          **{"crit__first_gen__eq__yes": T}),
        s("S055", "Martha H. Jones Endowed", 12920.22,
          **{"crit__school__contains__unr": T}),
        s("S056", "Meggin McIntosh Endowed", 3930.79,
          **{"crit__school__contains__unr": T, "crit__age__gte__25": T, "crit__age__lte__55": T}),
        s("S057", "Michele Attaway Endowed", 1000),   # domestic violence — no data
        s("S058", "Milton & Peggy Glick Endowed", 2002.89),
        s("S059", "Nancy Peppin Endowed", 2191.72),
        s("S060", "Nazir and Mary Ansari", 30000),
        s("S061", "Nevada Gold Mines", 2500,
          **{"crit__rural__eq__yes": T}),
        s("S062", "Northern Nevada Nursing Endowed", 51273.85,
          **{"crit__rural__eq__yes": T, "crit__major__contains__nursing": T}),
        s("S063", "NVEnergy", 2500,
          **{"crit__stem__eq__yes": T, "crit__rural__eq__yes": T}),
        s("S064", "NWF Board Alumni", 1000),
        s("S065", "NWF General", 25000),
        s("S066", "Ophelia Martinez", 2300),
        s("S067", "Orduna Hastings", 3000,
          **{"crit__major__contains__law": T}),
        s("S068", "Pam Dolan", 2500),
        s("S069", "Pauline Helms Endowed", 1022.21),   # over 40 + business/CS — complex
        s("S070", "Peek", 8000,
          **{"crit__rural__eq__yes": T}),
        s("S071", "Peg Murphy Endowed", 2500,
          **{"crit__single_mother__eq__yes": T}),   # partial: also community college
        s("S072", "Peggy Bowker Endowed", 7500),
        s("S073", "PK Electrical", 5000,
          **{"crit__major__contains__engineering": T}),
        s("S074", "PPG Endowed Scholarship", 4500,
          **{"crit__major__contains__accounting": T}),
        s("S075", "Raffealli Family Endowed", 7500,
          **{"crit__major__contains__education": T}),
        s("S076", "Redemption Fund - Jill Tolles", 10000),   # domestic violence — no data
        s("S077", "Redfield", 10000,
          **{"crit__tutoring__eq__yes": T, "crit__washoe_county__eq__yes": T}),
        s("S078", "Reno Gazette Journal Endowed", 1164.69),
        s("S079", "Renown", 2500),   # medicine/nursing/science/math — OR, general
        s("S080", "Ross and Dorothy Ballard nursing or medical", 3000),   # nursing OR medical, general
        s("S081", "Ruth Gaiser Brown", 2500,
          **{"crit__single_mother__eq__yes": T}),   # partial: also health/education
        s("S082", "Sandvik", 2000,
          **{"crit__rural__eq__yes": T}),
        s("S083", "Sandy McKee Endowed", 2500,
          **{"crit__first_gen__eq__yes": T}),
        s("S084", "Sarah Dean Gaiser Nursing", 2500,
          **{"crit__single_mother__eq__yes": T, "crit__major__contains__nursing": T}),
        s("S085", "Sher & Gary Muhonen Endowed", 2159.12,
          **{"crit__major__contains__engineering": T}),   # CS OR EE — engineering catches EE
        s("S086", "SilverSummit", 40000,
          **{"crit__silversummit__eq__yes": T}),
        s("S087", "Society of Women Engineers Endowed", 1053.88,
          **{"crit__major__contains__engineering": T}),
        s("S088", "Soroptimists", 30000,
          **{"crit__gap_in_studies__eq__yes": T}),   # re-entry
        s("S089", "State 36", 3000),
        s("S090", "Sue Wagner Endowed", 2940.88,
          **{"crit__school__contains__unr": T, "crit__stem__eq__yes": T}),
        s("S091", "Susanne S. Stout Endowed", 2253.23),
        s("S092", "Terry Lee Wells Endowed", 1395.59,
          **{"crit__rural__eq__yes": T, "crit__gap_in_studies__eq__yes": T}),
        s("S093", "Timken Sturgis", 1000,
          **{"crit__major__contains__medicine": T}),
        s("S094", "Turner", 4000,
          **{"crit__construction_field__eq__yes": T}),
        s("S095", "Wawona", 5000),
        s("S096", "Wild Blueberry Endowed", 2500,
          **{"crit__single_mother__eq__yes": T, "crit__stem__eq__yes": T}),
        s("S097", "WITS Endowed", 2729.56,
          **{"crit__major__contains__engineering": T}),   # partial: also Jr/Sr
        s("S098", "Women in Tesla", 5752.08,
          **{"crit__major__contains__engineering": T, "crit__grade__eq__freshman": T}),
        s("S099", "Wood Rodgers", 2500,
          **{"crit__stem__eq__yes": T}),
        # Reno Connection Network has $0 — skip
    ]


def build_manual_allocations(ws, recipients: list[dict], scholarships_by_col: dict) -> list[dict]:
    """Extract the client's actual manual allocation from the spreadsheet."""
    # Map recipient_id → Excel row index (1-based)
    rid_to_row: dict[str, int] = {}
    for excel_row in range(7, 243):
        num = ws.cell(excel_row, 2).value
        if num is None or str(num).strip() in ("", " "):
            continue
        rid = f"R{str(num).strip().zfill(3)}"
        rid_to_row[rid] = excel_row

    # Map scholarship col → scholarship name
    allocations = []
    for r in recipients:
        rid = r["recipient_id"]
        excel_row = rid_to_row.get(rid)
        if excel_row is None:
            continue
        for col, schol_info in scholarships_by_col.items():
            val = ws.cell(excel_row, col).value
            if val and isinstance(val, (int, float)) and val > 0:
                allocations.append({
                    "recipient_id": rid,
                    "scholarship_name": schol_info["name"],
                    "amount": round(float(val), 2),
                })
    return allocations


def main() -> None:
    OUT_DIR.mkdir(exist_ok=True)

    print(f"Loading {EXCEL_PATH}...")
    wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)
    ws = wb["Sheet1"]

    recipients, scholarships_by_col = build_recipients(ws)
    print(f"  Funded recipients: {len(recipients)}")
    print(f"  Total awards: ${sum(r['award_amount'] for r in recipients):,.2f}")

    scholarships = build_scholarships()
    total_pool = sum(s["amount"] for s in scholarships)
    print(f"  Scholarships: {len(scholarships)}")
    print(f"  Total pool: ${total_pool:,.2f}")

    # Write recipients CSV (drop internal _row field)
    recip_path = OUT_DIR / "recipients.csv"
    recip_clean = [{k: v for k, v in r.items() if k != "_row"} for r in recipients]
    with open(recip_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=recip_clean[0].keys())
        writer.writeheader()
        writer.writerows(recip_clean)
    print(f"  Written: {recip_path}")

    # Write scholarships CSV
    schol_path = OUT_DIR / "scholarships.csv"
    with open(schol_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=scholarships[0].keys())
        writer.writeheader()
        writer.writerows(scholarships)
    print(f"  Written: {schol_path}")

    # Write manual allocations CSV
    manual = build_manual_allocations(ws, recipients, scholarships_by_col)
    manual_path = OUT_DIR / "manual_allocations.csv"
    with open(manual_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["recipient_id", "scholarship_name", "amount"])
        writer.writeheader()
        writer.writerows(manual)
    manual_total = sum(m["amount"] for m in manual)
    print(f"  Manual allocations: {len(manual)} rows, total ${manual_total:,.2f}")
    print(f"  Written: {manual_path}")

    # Quick sanity check: flag recipients with no manual allocation
    manual_rids = {m["recipient_id"] for m in manual}
    missing = [r["recipient_id"] for r in recipients if r["recipient_id"] not in manual_rids]
    if missing:
        print(f"\n  WARNING: {len(missing)} funded recipients have no manual allocations: {missing}")


if __name__ == "__main__":
    main()
